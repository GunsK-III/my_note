# 在Excel中用单元格画画 (^_^)
"""具体说明等会再写……"""

"""第一步：先压缩图片"""
import cv2

image_path = input(r"输入图片的路径(路径前后不要用引号): ")
img = cv2.imread(image_path)
if img is None:
    print("路径输入有误！")
# cv2.imshow("Image", img)
# cv2.waitKey(0)
# cv2.destroyAllWindows()

width, height = img.shape[1], img.shape[0]
print("图片的宽度为：", width, "像素.", "图片的高度为：", height, "像素.")
print("下面分别输入图片压缩后的宽度和高度。"
      "为了最后显示效果美观，建议了解图片原尺寸后输入等比例的宽高数值")

new_width = int(input("输入图片横向占用单元格的数量(正整数)："))
new_height = int(input("输入图片纵向占用单元格的数量(正整数)："))
resized = cv2.resize(img, (new_width, new_height))
if_write = input("如果要保存压缩后的图片，输入“是”或“y”，否则键入“enter”：")
if if_write == "y" or if_write == "是":
    print("图片默认保存在与本文件相同路径下。")
    output_path = 'yasuohou.jpg'
    cv2.imwrite(output_path, resized)
    print("压缩图片已保存！")
else:
    pass

"""第二步：创建一个excel，并格式化其中中的单元格"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # 这个模块针不戳


def create_and_modify_excel(output_path):
    wb = Workbook()
    ws = wb.active
    row_height = float(input("请输入行高的数值(建议填入10): "))
    column_width = float(input("请输入列宽数值(建议填入2): "))
    for row in range(1, new_height + 1):  # 修改行高
        ws.row_dimensions[row].height = row_height
    for col in range(1, new_width + 1):  # 修改列宽
        ws.column_dimensions[get_column_letter(col)].width = column_width
    wb.save(output_path)


output_path = 'Hi~,^_^.xlsx'
create_and_modify_excel(output_path)

"""第三步：获取图片像素颜色，并填入到表格中。"""
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

image = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)  # 将BGR转换为RGB
workbook = load_workbook('Hi~,^_^.xlsx')
sheet = workbook.active

for i in range(new_height):  # 行
    for j in range(new_width):  # 列
        r, g, b = image[i, j]
        color = f"FF{r:02X}{g:02X}{b:02X}"
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet.cell(row=i + 1, column=j + 1).fill = fill

workbook.save('Hi~,^_^.xlsx')
input('打开 "Hi~, ^_^" excel表格看看吧，位置和在本文件在同一个路径下。'
      '程序已结束，按下Enter键退出。')
